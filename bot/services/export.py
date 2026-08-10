"""Выгрузка заказов в XLSX."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from bot.db.models import Order, OrderStatus, User

HEADERS = [
    ("ID заказа", 12),
    ("Дата создания", 19),
    ("Дата оплаты", 19),
    ("Telegram ID", 14),
    ("Username", 18),
    ("Тип", 12),
    ("Товар", 40),
    ("Кол-во", 8),
    ("Сумма до скидки, ₽", 18),
    ("Скидка, ₽", 12),
    ("Итого, ₽", 12),
    ("Промокод", 16),
    ("Способ оплаты", 18),
    ("ID транзакции", 38),
    ("Статус", 16),
]


def _rub(kop: int | None) -> float:
    return round(int(kop or 0) / 100, 2)


def _dt(value: datetime | None) -> str:
    return value.strftime("%Y-%m-%d %H:%M") if value else ""


async def orders_workbook(
    session: AsyncSession,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    status: str | None = None,
) -> tuple[bytes, int]:
    """Собирает книгу заказов. Возвращает `(содержимое, число строк)`."""
    stmt = select(Order, User).outerjoin(User, User.tg_id == Order.user_id)
    if date_from:
        stmt = stmt.where(Order.created_at >= date_from)
    if date_to:
        stmt = stmt.where(Order.created_at <= date_to)
    if status:
        stmt = stmt.where(Order.status == status)
    stmt = stmt.order_by(Order.created_at.desc())

    rows = (await session.execute(stmt)).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Заказы"

    for column, (title, width) in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=column, value=title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A2"

    for index, (order, user) in enumerate(rows, start=2):
        sheet.cell(row=index, column=1, value=order.id)
        sheet.cell(row=index, column=2, value=_dt(order.created_at))
        sheet.cell(row=index, column=3, value=_dt(order.paid_at))
        sheet.cell(row=index, column=4, value=order.user_id)
        sheet.cell(row=index, column=5, value=(user.username if user else "") or "")
        sheet.cell(row=index, column=6, value=order.kind)
        sheet.cell(row=index, column=7, value=order.product_title)
        sheet.cell(row=index, column=8, value=order.qty)
        sheet.cell(row=index, column=9, value=_rub(order.subtotal_kop))
        sheet.cell(row=index, column=10, value=_rub(order.discount_kop))
        sheet.cell(row=index, column=11, value=_rub(order.total_kop))
        sheet.cell(row=index, column=12, value=order.promo_code or "")
        sheet.cell(row=index, column=13, value=order.payment_method or "")
        sheet.cell(row=index, column=14, value=order.provider_txn_id or "")
        sheet.cell(
            row=index,
            column=15,
            value=OrderStatus.TITLES.get(order.status, order.status),
        )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), len(rows)
